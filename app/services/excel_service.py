from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Dict, List, Optional

from openpyxl import load_workbook

from app.models import SampleTagData


HEADER_ALIASES = {
    "STYLE NO": "style_no",
    "COLOR CODE": "color_code",
    "FACTORY": "factory",
    "INITIAL DELIVERY (SHIP DATE)": "ship_date",
    "FABRIC INFO": "fabric_info",
    "FNF DESIGNER": "designer",
    "FNF TD": "td",
}


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split()).strip().upper()


def excel_serial_to_date(value: object) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        base = datetime(1899, 12, 30)
        try:
            return (base + timedelta(days=float(value))).strftime("%Y-%m-%d")
        except Exception:
            return str(value)
    return str(value).strip()


def default_size(style_no: str) -> str:
    return "S" if (style_no or "").upper().startswith("3F") else "L"


def normalize_factory(value: object) -> str:
    return "" if value is None else str(value).strip()


@dataclass
class RawDataStore:
    styles: Dict[str, SampleTagData]

    @classmethod
    def empty(cls) -> "RawDataStore":
        return cls(styles={})

    @classmethod
    def from_xlsx_bytes(cls, content: bytes) -> "RawDataStore":
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]

        header_row = None
        col_map: Dict[str, int] = {}
        for row_idx in range(1, min(ws.max_row, 20) + 1):
            current = {}
            for col_idx, cell in enumerate(ws[row_idx], start=1):
                normalized = normalize_header(cell.value)
                if normalized in HEADER_ALIASES:
                    current[HEADER_ALIASES[normalized]] = col_idx
            if "style_no" in current:
                header_row = row_idx
                col_map = current
                break

        if header_row is None:
            raise ValueError("Could not find STYLE NO header in workbook")

        store: Dict[str, SampleTagData] = {}
        conflicts: Dict[str, set] = {}

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            def get(field: str):
                idx = col_map.get(field)
                return row[idx - 1] if idx else None

            style = str(get("style_no") or "").strip()
            if not style:
                continue

            candidate = SampleTagData(
                style_sample_no=style,
                size=default_size(style),
                description=str(get("color_code") or "").strip(),
                vendor="Nobland",
                factory=normalize_factory(get("factory")),
                fabric_info=str(get("fabric_info") or "").strip(),
                e_pattern="YES",
                ship_date=excel_serial_to_date(get("ship_date")),
                designer=str(get("designer") or "").strip(),
                td=str(get("td") or "").strip(),
            )

            if style not in store:
                store[style] = candidate
            else:
                existing = store[style]
                for attr in ["description", "factory", "fabric_info", "ship_date", "designer", "td"]:
                    new_value = getattr(candidate, attr)
                    old_value = getattr(existing, attr)
                    if not old_value and new_value:
                        setattr(existing, attr, new_value)
                    elif new_value and old_value and new_value != old_value:
                        conflicts.setdefault(style, set()).add(attr)

        wb.close()
        return cls(styles=store)

    def list_styles(self, query: str = "") -> List[str]:
        query = (query or "").strip().upper()
        items = sorted(self.styles.keys())
        if not query:
            return items
        return [s for s in items if query in s.upper()]

    def get(self, style_no: str) -> Optional[SampleTagData]:
        return self.styles.get(style_no)
